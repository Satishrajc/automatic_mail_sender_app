# Python code to illustrate Sending mail with attachments

# libraries to be imported
import os
import sys
import openpyxl
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

# Please give the bellow info/paths which required to run this script
your_mail_id = "your_mailid"
your_password = "your_password"
excel_data_path = os.path.join(os.getcwd(), 'mail_data.xlsx')
attachment_file_path = os.path.join(os.getcwd(), 'Resume_Satish_Python_4_YrsExp.docx')
mail_body_path = os.path.join(os.getcwd(), 'mail_body.txt')
header_text = "[Resume] Satish - Python Automation – 4+ Years of Experience"


class MailSending(object):
    def __init__(self, mail_id, password, attached_file, header):
        """ Initialisation """
        self.your_mail_id = mail_id
        self.your_password = password
        self.attachment_file_path = attached_file
        self.header_text = header

    def read_mail_body(self, mail_body_path):
        """ Reading the mail body text from the given path/file"""
        try:
            fd = open(mail_body_path)
            self.mail_body = fd.read()

        except IOError:
            print('Given mail_body_path is wrong: {}'.format(mail_body_path))
            sys.exit(-1)

        except Exception:
            print('Error: {} - Got error in method: read_mail_body()'.format(sys.exc_info()[1:-1]))

    def get_data_from_excel(self, excel_sheet_path):
        """ Reading names and email ids from the given path/excel file """
        try:
            wb = openpyxl.load_workbook(excel_sheet_path)
            sheet = wb.get_sheet_by_name('name_and_mailID')

            dict_sending_info = {}
            for roww in range(2, sheet.max_row + 1):
                dict_sending_info[sheet.cell(row=roww, column=2).value] = sheet.cell(row=roww, column=1).value
            return dict_sending_info

        except IOError:
            print('Given excel_data_path is wrong: {}'.format(excel_sheet_path))
            sys.exit(-1)

        except Exception:
            print('Error: {} - Got error in method: get_data_from_excel()'.format(sys.exc_info()[1:-1]))

    def send_mail(self, name, to_address):
        """ Establishing the SMTP server connection and sending the mail"""
        try:
            attachment = open(self.attachment_file_path, "rb")
            attached_filename = os.path.basename(self.attachment_file_path)
            msg = MIMEMultipart()
            msg['Subject'] = self.header_text
            msg['From'] = self.your_mail_id
            msg['To'] = to_address

            body = "Dear {0}, \n\n".format(name) + self.mail_body
            msg.attach(MIMEText(body, 'plain'))

            p = MIMEBase('application', 'octet-stream')
            p.set_payload((attachment).read())
            encoders.encode_base64(p)
            p.add_header('Content-Disposition', "attachment; filename= %s" % attached_filename)
            msg.attach(p)

            server = smtplib.SMTP('smtp.gmail.com', 587)
            server.starttls()
            server.login(self.your_mail_id, self.your_password)
            text = msg.as_string()
            server.sendmail(self.your_mail_id, to_address, text)
            server.quit()

        except IOError:
            print('Given attachment_file_path is wrong: {}'.format(self.attachment_file_path))
            sys.exit(-1)

        except Exception:
            print('Error: {} - Got error in method: send_mail()'.format(sys.exc_info()[1:-1]))


def main():
    """ Triggering all the functions/methods """
    try:
        # Creating the object
        object_send_mail = MailSending(your_mail_id, your_password, attachment_file_path, header_text)

        object_send_mail.read_mail_body(mail_body_path)
        data_name_mail_id = object_send_mail.get_data_from_excel(excel_data_path)

        if data_name_mail_id:
            for mail_id, name in data_name_mail_id.items():
                print("Name: {0} - Mail ID: {1}".format(name, mail_id))
                object_send_mail.send_mail(name, mail_id)
        else:
            print('Could not read from excel sheet')
            sys.exit(-1)

    except Exception:
        print('Error: {} - Got error in function: main()'.format(sys.exc_info()[1:-1]))
        sys.exit(-1)

    else:
        sys.exit(0)


if __name__ == "__main__":
    main()